#!/usr/bin/env python3
"""
CSV to XLSX Converter with Conditional Formatting for RISC-V Extensions

This script converts a CSV file containing RISC-V extension support data
to an Excel XLSX file with conditional formatting (Y=green, N=red).
"""

import argparse
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Convert RISC-V extensions CSV to XLSX with conditional formatting'
    )
    parser.add_argument('input_csv', help='Input CSV file path')
    parser.add_argument(
        '--output', '-o', 
        help='Output XLSX file path (default: same filename with .xlsx extension)'
    )
    parser.add_argument(
        '--freeze', '-f', action='store_true',
        help='Freeze the first two columns (Name, Version)'
    )
    parser.add_argument(
        '--separate-sheets', '-s', action='store_true',
        help='Create separate sheets for GCC and Clang'
    )
    
    return parser.parse_args()

def apply_conditional_formatting(worksheet, start_col, data_rows):
    """Apply conditional formatting to Y/N cells."""
    # Define fills for Y and N values
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Get the last column letter
    last_col = openpyxl.utils.get_column_letter(worksheet.max_column)
    first_data_row = 2  # Header is row 1
    
    # Apply conditional formatting rules to the data range
    data_range = f"{start_col}{first_data_row}:{last_col}{data_rows + first_data_row - 1}"
    
    # Add the rules
    worksheet.conditional_formatting.add(
        data_range,
        CellIsRule(operator="equal", formula=['"Y"'], fill=green_fill)
    )
    worksheet.conditional_formatting.add(
        data_range,
        CellIsRule(operator="equal", formula=['"N"'], fill=red_fill)
    )

def format_worksheet(worksheet, freeze=True, has_description=True):
    """Format the worksheet with styling and column widths."""
    # Set column widths
    worksheet.column_dimensions['A'].width = 15  # Name
    worksheet.column_dimensions['B'].width = 8   # Version
    
    if has_description:
        worksheet.column_dimensions['C'].width = 40  # Description
    
    # Set header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Freeze panes if requested
    if freeze:
        # If we have description, freeze first 3 columns, otherwise first 2
        freeze_col = 'D' if has_description else 'C'
        worksheet.freeze_panes = f"{freeze_col}2"

def create_separate_sheets(df, output_file, freeze):
    """Create separate sheets for GCC and Clang."""
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Get list of all compiler columns
        all_cols = list(df.columns)
        
        # Base columns that always stay
        base_cols = ['Name', 'Version']
        if 'Description' in all_cols:
            base_cols.append('Description')
            has_description = True
        else:
            has_description = False
        
        # Filter for GCC and Clang columns
        gcc_cols = [col for col in all_cols if col.startswith('gcc-')]
        clang_cols = [col for col in all_cols if col.startswith('clang-')]
        
        # Create combined sheet with all data
        df.to_excel(writer, sheet_name='All Extensions', index=False)
        
        # Create GCC sheet
        if gcc_cols:
            # Combine base columns with GCC columns
            gcc_df = df[base_cols + gcc_cols]
            gcc_df.to_excel(writer, sheet_name='GCC Extensions', index=False)
        
        # Create Clang sheet
        if clang_cols:
            # Combine base columns with Clang columns
            clang_df = df[base_cols + clang_cols]
            clang_df.to_excel(writer, sheet_name='Clang Extensions', index=False)
        
        # Format all sheets after writing
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Format the worksheet
            format_worksheet(worksheet, freeze, has_description)
            
            # Apply conditional formatting
            data_start_col = 'D' if has_description else 'C'
            apply_conditional_formatting(worksheet, data_start_col, len(df))

def main():
    """Main function."""
    args = parse_arguments()
    
    # Set output file if not specified
    if not args.output:
        base_name = os.path.splitext(args.input_csv)[0]
        args.output = f"{base_name}.xlsx"
    
    # Load CSV data
    df = pd.read_csv(args.input_csv)
    
    print(f"Converting {args.input_csv} to {args.output}")
    print(f"Found {len(df)} extensions across {len(df.columns) - 2} compiler versions")
    
    if args.separate_sheets:
        create_separate_sheets(df, args.output, args.freeze)
    else:
        # If not creating separate sheets, just format the main sheet
        with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            # Check if Description column exists
            has_description = 'Description' in df.columns
            
            # Format the worksheet
            format_worksheet(worksheet, args.freeze, has_description)
            
            # Apply conditional formatting starting from the first data column
            data_start_col = 'D' if has_description else 'C'
            apply_conditional_formatting(worksheet, data_start_col, len(df))
    
    print(f"✅ Successfully created {args.output}")
    
    # Print examples of how to use
    print("\nUsage examples:")
    print(f"  Basic conversion:            python csv_to_xlsx.py {args.input_csv}")
    print(f"  With separate sheets:        python csv_to_xlsx.py {args.input_csv} --separate-sheets")
    print(f"  Without freezing columns:    python csv_to_xlsx.py {args.input_csv} --freeze")
    print(f"  Custom output:               python csv_to_xlsx.py {args.input_csv} -o custom_output.xlsx")

if __name__ == "__main__":
    main()
