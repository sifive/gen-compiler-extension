#!/usr/bin/env python3

import os
import sys
import pandas as pd
import argparse
from pathlib import Path

def convert_csv_to_excel(input_file, output_file=None, sheet_name=None):
    """
    Convert a CSV file to Excel format.
    
    Args:
        input_file (str): Path to the input CSV file
        output_file (str, optional): Path to the output Excel file. If not provided,
                                    uses the same name as input with .xlsx extension
        sheet_name (str, optional): Name of the sheet in Excel. If not provided,
                                   uses the basename of the file without extension
    
    Returns:
        str: Path to the created Excel file
    """
    # Determine output file name if not provided
    if not output_file:
        output_file = str(Path(input_file).with_suffix('.xlsx'))
    
    # Determine sheet name if not provided
    if not sheet_name:
        sheet_name = Path(input_file).stem
    
    print(f"Converting {input_file} to {output_file} (sheet: {sheet_name})")
    
    # Read the CSV file
    df = pd.read_csv(input_file)
    
    # Write to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"Conversion complete: {output_file}")
    return output_file

def convert_directory(input_dir, output_dir=None, recursive=False):
    """
    Convert all CSV files in a directory to Excel format.
    
    Args:
        input_dir (str): Path to the input directory containing CSV files
        output_dir (str, optional): Path to the output directory for Excel files
        recursive (bool): Whether to process subdirectories recursively
    """
    input_path = Path(input_dir)
    
    # Determine output directory
    if output_dir:
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
    else:
        output_path = input_path
    
    # Get all CSV files
    if recursive:
        csv_files = list(input_path.glob('**/*.csv'))
    else:
        csv_files = list(input_path.glob('*.csv'))
    
    if not csv_files:
        print(f"No CSV files found in {input_dir}")
        return
    
    print(f"Found {len(csv_files)} CSV files to convert")
    
    # Convert each file
    for csv_file in csv_files:
        # Determine relative path from input directory
        rel_path = csv_file.relative_to(input_path)
        
        # Create output directory structure if needed
        output_file_dir = output_path / rel_path.parent
        output_file_dir.mkdir(parents=True, exist_ok=True)
        
        # Determine output file path
        output_file = output_file_dir / f"{rel_path.stem}.xlsx"
        
        # Convert the file
        convert_csv_to_excel(str(csv_file), str(output_file))

def convert_version_comparison(input_file, output_file=None):
    """
    Convert the version comparison CSV to a formatted Excel file with conditional formatting.
    
    Args:
        input_file (str): Path to the version comparison CSV file
        output_file (str, optional): Path to the output Excel file
    
    Returns:
        str: Path to the created Excel file
    """
    # Determine output file name if not provided
    if not output_file:
        output_file = str(Path(input_file).with_suffix('.xlsx'))
    
    print(f"Converting version comparison {input_file} to {output_file}")
    
    # Read the CSV file
    df = pd.read_csv(input_file)
    
    # Write to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Version Comparison', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Version Comparison']
        
        # Add conditional formatting (highlight cells with 'X')
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import FormulaRule
        
        # Light green fill for cells with 'X'
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        # Apply to all data cells
        for col in range(2, len(df.columns) + 1):  # Start from column B (index 2 in openpyxl)
            col_letter = worksheet.cell(row=1, column=col).column_letter
            worksheet.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df) + 1}",
                FormulaRule(formula=[f'${col_letter}2="X"'], fill=green_fill)
            )
        
        # Freeze the first row and column
        worksheet.freeze_panes = 'B2'
        
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
    
    print(f"Conversion complete: {output_file}")
    return output_file

def main():
    parser = argparse.ArgumentParser(description='Convert CSV files to Excel format')
    parser.add_argument('input', help='Input CSV file or directory')
    parser.add_argument('-o', '--output', help='Output Excel file or directory')
    parser.add_argument('-r', '--recursive', action='store_true', help='Process directories recursively')
    parser.add_argument('-c', '--comparison', action='store_true', help='Format as version comparison file')
    
    args = parser.parse_args()
    
    input_path = Path(args.input)
    
    if not input_path.exists():
        print(f"Error: Input path {args.input} does not exist")
        sys.exit(1)
    
    if input_path.is_dir():
        convert_directory(args.input, args.output, args.recursive)
    else:
        if args.comparison:
            convert_version_comparison(args.input, args.output)
        else:
            convert_csv_to_excel(args.input, args.output)

if __name__ == '__main__':
    main()
